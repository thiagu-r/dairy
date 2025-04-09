import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, F, DecimalField, Count, Q, Max
from django.db.models.functions import Coalesce, TruncDay, TruncWeek, TruncMonth, TruncYear
from decimal import Decimal

from apps.seller.models import Seller
from apps.sales.models import SalesOrder, OrderItem
from apps.delivery.models import DeliveryOrder
from apps.products.models import Product

def get_date_trunc_func(time_filter):
    """Return the appropriate date truncation function based on the time filter."""
    trunc_funcs = {
        'daily': TruncDay,
        'weekly': TruncWeek,
        'monthly': TruncMonth,
        'yearly': TruncYear,
    }
    return trunc_funcs.get(time_filter, TruncDay)

def get_sales_analytics_by_seller(start_date=None, end_date=None, time_filter='monthly'):
    """Get sales analytics data grouped by seller."""
    queryset = OrderItem.objects.select_related('order', 'order__seller', 'product')

    # Apply date filters if provided
    if start_date:
        queryset = queryset.filter(order__delivery_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(order__delivery_date__lte=end_date)

    # Group by seller and aggregate
    analytics = queryset.values(
        'order__seller'
    ).annotate(
        seller_id=F('order__seller__id'),
        seller__store_name=F('order__seller__store_name'),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('unit_price'), output_field=DecimalField()),
        order_count=Count('order', distinct=True)
    ).order_by('-total_value')

    return analytics

def get_sales_analytics_by_product(start_date=None, end_date=None, time_filter='monthly'):
    """Get sales analytics data grouped by product."""
    queryset = OrderItem.objects.select_related('product', 'order')

    # Apply date filters if provided
    if start_date:
        queryset = queryset.filter(order__delivery_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(order__delivery_date__lte=end_date)

    # Group by product and aggregate
    analytics = queryset.values(
        'product'
    ).annotate(
        product_id=F('product__id'),
        product__name=F('product__name'),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('unit_price'), output_field=DecimalField()),
        order_count=Count('order', distinct=True)
    ).order_by('-total_value')

    return analytics

def get_sales_analytics_by_route(start_date=None, end_date=None, time_filter='monthly'):
    """Get sales analytics data grouped by route."""
    queryset = OrderItem.objects.select_related('order', 'order__seller', 'order__seller__route')

    # Apply date filters if provided
    if start_date:
        queryset = queryset.filter(order__delivery_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(order__delivery_date__lte=end_date)

    # Group by route and aggregate
    analytics = queryset.values(
        'order__seller__route'
    ).annotate(
        route_id=F('order__seller__route__id'),
        route__name=F('order__seller__route__name'),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('unit_price'), output_field=DecimalField()),
        order_count=Count('order', distinct=True)
    ).order_by('-total_value')

    return analytics

def get_seller_balance_report(seller_id=None, start_date=None, end_date=None):
    """Get seller balance report data."""
    queryset = Seller.objects.all()

    # Filter by seller if provided
    if seller_id:
        queryset = queryset.filter(id=seller_id)

    # Prepare date filters for related queries
    date_filter_delivery = Q()
    date_filter_payments = Q()

    if start_date:
        date_filter_delivery &= Q(delivery_orders__delivery_date__gte=start_date)
        date_filter_payments &= Q(delivery_orders__payment_date__gte=start_date)

    if end_date:
        date_filter_delivery &= Q(delivery_orders__delivery_date__lte=end_date)
        date_filter_payments &= Q(delivery_orders__payment_date__lte=end_date)

    # Annotate with aggregated data
    sellers = queryset.annotate(
        opening_balance=Coalesce(Sum('delivery_orders__opening_balance', filter=date_filter_delivery), Decimal('0.00')),
        total_sales=Coalesce(Sum('delivery_orders__total_price', filter=date_filter_delivery), Decimal('0.00')),
        total_payments=Coalesce(Sum('delivery_orders__amount_collected', filter=date_filter_payments), Decimal('0.00')),
        current_balance=F('opening_balance') + F('total_sales') - F('total_payments'),
        last_payment_date=Max('delivery_orders__delivery_date', filter=Q(delivery_orders__amount_collected__gt=0)),
        last_order_date=Max('delivery_orders__delivery_date')
    ).order_by('store_name')

    return sellers

def generate_excel_for_sales(queryset, group_by='seller', start_date=None, end_date=None):
    """Generate Excel file for sales data."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set title based on grouping
    title_mapping = {
        'seller': 'Sales by Seller',
        'product': 'Sales by Product',
        'route': 'Sales by Route'
    }
    worksheet.title = title_mapping.get(group_by, 'Sales Report')

    # Add report header
    worksheet.merge_cells('A1:F1')
    header_cell = worksheet['A1']
    header_cell.value = f"{worksheet.title} - {start_date} to {end_date}" if start_date and end_date else worksheet.title
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')

    # Define column headers based on grouping
    if group_by == 'seller':
        headers = ['Seller ID', 'Seller Name', 'Total Quantity', 'Total Value', 'Order Count']
    elif group_by == 'product':
        headers = ['Product ID', 'Product Name', 'Total Quantity', 'Total Value', 'Order Count']
    elif group_by == 'route':
        headers = ['Route ID', 'Route Name', 'Total Quantity', 'Total Value', 'Order Count']
    else:
        headers = ['ID', 'Name', 'Total Quantity', 'Total Value', 'Order Count']

    # Add column headers
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for row_num, item in enumerate(queryset, 4):
        if group_by == 'seller':
            worksheet.cell(row=row_num, column=1).value = item['seller_id']
            worksheet.cell(row=row_num, column=2).value = item['seller__store_name']
        elif group_by == 'product':
            worksheet.cell(row=row_num, column=1).value = item['product_id']
            worksheet.cell(row=row_num, column=2).value = item['product__name']
        elif group_by == 'route':
            worksheet.cell(row=row_num, column=1).value = item['route_id']
            worksheet.cell(row=row_num, column=2).value = item['route__name']

        worksheet.cell(row=row_num, column=3).value = float(item['total_quantity'])
        worksheet.cell(row=row_num, column=4).value = float(item['total_value'])
        worksheet.cell(row=row_num, column=5).value = item['order_count']

    # Add totals row
    total_row = row_num + 2
    worksheet.cell(row=total_row, column=1).value = "TOTAL"
    worksheet.cell(row=total_row, column=1).font = Font(bold=True)

    # Calculate totals
    total_quantity = sum(float(item['total_quantity']) for item in queryset)
    total_value = sum(float(item['total_value']) for item in queryset)
    total_orders = sum(item['order_count'] for item in queryset)

    worksheet.cell(row=total_row, column=3).value = total_quantity
    worksheet.cell(row=total_row, column=3).font = Font(bold=True)
    worksheet.cell(row=total_row, column=4).value = total_value
    worksheet.cell(row=total_row, column=4).font = Font(bold=True)
    worksheet.cell(row=total_row, column=5).value = total_orders
    worksheet.cell(row=total_row, column=5).font = Font(bold=True)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    return output

def generate_excel_for_seller_balance(queryset, start_date=None, end_date=None):
    """Generate Excel file for seller balance report."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Seller Balance Report'

    # Add report header
    worksheet.merge_cells('A1:G1')
    header_cell = worksheet['A1']
    header_cell.value = f"Seller Balance Report - {start_date} to {end_date}" if start_date and end_date else "Seller Balance Report"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')

    # Define column headers
    headers = ['Seller ID', 'Store Name', 'Owner Name', 'Opening Balance', 'Total Sales',
               'Total Payments', 'Current Balance', 'Last Payment Date', 'Last Order Date']

    # Add column headers
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for row_num, seller in enumerate(queryset, 4):
        worksheet.cell(row=row_num, column=1).value = seller.id
        worksheet.cell(row=row_num, column=2).value = seller.store_name
        worksheet.cell(row=row_num, column=3).value = seller.owner_name
        worksheet.cell(row=row_num, column=4).value = float(seller.opening_balance)
        worksheet.cell(row=row_num, column=5).value = float(seller.total_sales)
        worksheet.cell(row=row_num, column=6).value = float(seller.total_payments)
        worksheet.cell(row=row_num, column=7).value = float(seller.current_balance)
        worksheet.cell(row=row_num, column=8).value = seller.last_payment_date
        worksheet.cell(row=row_num, column=9).value = seller.last_order_date

        # Highlight negative balances
        if seller.current_balance < 0:
            for col in range(1, 10):
                worksheet.cell(row=row_num, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Add totals row
    total_row = row_num + 2
    worksheet.cell(row=total_row, column=1).value = "TOTAL"
    worksheet.cell(row=total_row, column=1).font = Font(bold=True)

    # Calculate totals
    total_opening = sum(seller.opening_balance for seller in queryset)
    total_sales = sum(seller.total_sales for seller in queryset)
    total_payments = sum(seller.total_payments for seller in queryset)
    total_balance = sum(seller.current_balance for seller in queryset)

    worksheet.cell(row=total_row, column=4).value = float(total_opening)
    worksheet.cell(row=total_row, column=4).font = Font(bold=True)
    worksheet.cell(row=total_row, column=5).value = float(total_sales)
    worksheet.cell(row=total_row, column=5).font = Font(bold=True)
    worksheet.cell(row=total_row, column=6).value = float(total_payments)
    worksheet.cell(row=total_row, column=6).font = Font(bold=True)
    worksheet.cell(row=total_row, column=7).value = float(total_balance)
    worksheet.cell(row=total_row, column=7).font = Font(bold=True)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    return output

def generate_excel_for_detailed_sales(queryset, start_date=None, end_date=None):
    """Generate Excel file for detailed sales data including order items."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detailed Sales Report'

    # Add report header
    worksheet.merge_cells('A1:H1')
    header_cell = worksheet['A1']
    header_cell.value = f"Detailed Sales Report - {start_date} to {end_date}" if start_date and end_date else "Detailed Sales Report"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='center')

    # Define column headers
    headers = ['Order Number', 'Seller Name', 'Order Date', 'Product Name', 'Product Code',
               'Quantity', 'Unit Price', 'Total Price']

    # Add column headers
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for row_num, item in enumerate(queryset, 4):
        worksheet.cell(row=row_num, column=1).value = item.order.order_number
        worksheet.cell(row=row_num, column=2).value = item.order.seller.store_name
        worksheet.cell(row=row_num, column=3).value = item.order.delivery_date
        worksheet.cell(row=row_num, column=4).value = item.product.name
        worksheet.cell(row=row_num, column=5).value = item.product.code
        worksheet.cell(row=row_num, column=6).value = float(item.quantity)
        worksheet.cell(row=row_num, column=7).value = float(item.unit_price)
        worksheet.cell(row=row_num, column=8).value = float(item.quantity * item.unit_price)

    # Add totals row
    total_row = row_num + 2
    worksheet.cell(row=total_row, column=1).value = "TOTAL"
    worksheet.cell(row=total_row, column=1).font = Font(bold=True)

    # Calculate totals
    total_quantity = sum(item.quantity for item in queryset)
    total_price = sum(item.quantity * item.unit_price for item in queryset)

    worksheet.cell(row=total_row, column=6).value = float(total_quantity)
    worksheet.cell(row=total_row, column=6).font = Font(bold=True)
    worksheet.cell(row=total_row, column=8).value = float(total_price)
    worksheet.cell(row=total_row, column=8).font = Font(bold=True)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    return output
